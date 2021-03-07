import openpyxl


class ReportDataExtractor:
    def __init__(self, filename, acc_standards):
        self.results = list()
        acc_standards_row_mapping = dict()
        wb_report = openpyxl.load_workbook(filename)
        report_sheet = wb_report.active
        misses = 0
        current_standard_row = 7
        all_rows = acc_standards + ["Period End Date"]
        while misses < 3:
            cell_value = report_sheet.cell(current_standard_row, 2).value
            if not cell_value:
                current_standard_row += 1
                misses += 1
                continue
            misses = 0
            for st in all_rows:
                if st == cell_value:
                    acc_standards_row_mapping[st] = current_standard_row
                    break
            current_standard_row += 1

        period_date_column = 3
        period_date_row = acc_standards_row_mapping["Period End Date"]
        while True:
            end_date = report_sheet.cell(period_date_row, period_date_column).value
            if not end_date:
                break
            end_date = end_date[:11]
            values_for_date = dict()
            for k, v in acc_standards_row_mapping.items():
                value = report_sheet.cell(v, period_date_column).value
                if value and value != "--":
                    values_for_date[k] = value
            self.results.append({"date": end_date, "values": values_for_date})
            period_date_column += 1

    def get_json_data(self):
        return self.results