from datetime import datetime
import openpyxl


class QuotesDataExtractor:
    def __init__(self, filename):
        self.results = list()
        wb_quotes = openpyxl.load_workbook(filename)
        quotes_sheet = wb_quotes.active
        quote_row = 5
        misses = 0
        quotes_started = False
        while misses <= 3:
            value = quotes_sheet.cell(quote_row, 1).value
            if not value:
                misses += 1
                quote_row += 1
                continue
            elif value == "Exchange Date":
                quotes_started = True
            elif quotes_started:
                quote_date = value
                close = quotes_sheet.cell(quote_row, 2).value
                self.results.append({"date": quote_date.strftime("%d-%b-%Y"), "close": close})
           #     self.results.append({"date": quote_date.strftime("%d-%b-%Y"), "close": close})
            misses = 0
            quote_row += 1
        self.results.reverse()


    def get_json_data(self):
        return self.results

